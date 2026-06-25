import os
import sys
import json
from PIL import Image
from pypdf import PdfWriter
from concurrent.futures import ThreadPoolExecutor

base_dir = os.path.dirname(os.path.abspath(__file__))  # path absolut

def process_file(index, file_rel_path):
    """
    Worker function to convert a single file to PDF.
    Returns (index, pdf_path, is_temp, log_message)
    """
    full_path = os.path.join(base_dir, file_rel_path)
    if not os.path.exists(full_path):
        return index, None, False, f"File tidak ditemukan: {full_path}"
        
    ext = file_rel_path.lower().split('.')[-1]
    temp_pdf = full_path + ".temp.pdf"
    
    if ext == 'pdf':
        return index, full_path, False, f"Ditambahkan: {file_rel_path} (PDF)"
        
    elif ext in ['png', 'jpg', 'jpeg']:
        try:
            # Kompresi gambar: resize jika dimensi (lebar atau tinggi) melebihi 1200px
            img = Image.open(full_path)
            width, height = img.size
            max_size = 1200
            
            if width > max_size or height > max_size:
                if width > height:
                    new_width = max_size
                    new_height = int((max_size / width) * height)
                else:
                    new_height = max_size
                    new_width = int((max_size / height) * width)
                
                # Gunakan filter pencuplikan ulang LANCZOS untuk hasil terbaik
                try:
                    resample_filter = Image.Resampling.LANCZOS
                except AttributeError:
                    resample_filter = Image.LANCZOS
                
                img = img.resize((new_width, new_height), resample_filter)
                log_suffix = " (Compressed & Resized)"
            else:
                log_suffix = ""

            img = img.convert('RGB')
            img.save(temp_pdf)
            return index, temp_pdf, True, f"Converted & Ditambahkan{log_suffix}: {file_rel_path} -> Gambar"
        except Exception as e:
            return index, None, False, f"Gagal mengonversi Gambar {file_rel_path}: {str(e)}"
            
    elif ext == 'docx':
        try:
            import mammoth
            from xhtml2pdf import pisa
            
            # 1. Convert DOCX to HTML
            with open(full_path, "rb") as docx_file:
                result = mammoth.convert_to_html(docx_file)
                html_content = result.value
            
            # Simple, clean default styling for the converted Word document
            styled_html = f"""
            <html>
            <head>
            <meta http-equiv="Content-Type" content="text/html; charset=utf-8">
            <style>
                body {{
                    font-family: Helvetica, Arial, sans-serif;
                    font-size: 11pt;
                    line-height: 1.5;
                    color: #333333;
                }}
                h1, h2, h3, h4, h5, h6 {{
                    font-family: Helvetica, Arial, sans-serif;
                    color: #111111;
                    margin-top: 0.5cm;
                    margin-bottom: 0.3cm;
                }}
                p {{
                    margin-bottom: 0.3cm;
                    text-align: justify;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 0.5cm;
                    margin-bottom: 0.5cm;
                }}
                th, td {{
                    border: 1px solid #cccccc;
                    padding: 8px;
                    text-align: left;
                    font-size: 10pt;
                }}
                th {{
                    background-color: #f2f2f2;
                    font-weight: bold;
                }}
            </style>
            </head>
            <body>
                {html_content}
            </body>
            </html>
            """
            
            # 2. Convert HTML to PDF
            with open(temp_pdf, "wb") as result_file:
                pisa_status = pisa.CreatePDF(styled_html, dest=result_file)
                
            if pisa_status.err:
                raise Exception(f"xhtml2pdf error code: {pisa_status.err}")
                
            return index, temp_pdf, True, f"Converted & Ditambahkan: {file_rel_path} -> Word DOCX"
        except Exception as e:
            return index, None, False, f"Gagal mengonversi Word DOCX {file_rel_path}: {str(e)}"
            
    elif ext == 'txt':
        try:
            from fpdf import FPDF
            
            pdf = FPDF()
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            pdf.set_font("courier", size=10)
            
            with open(full_path, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    # Bersihkan karakter non-latin-1 agar tidak merusak font standar fpdf2
                    clean_line = line.encode('latin-1', 'replace').decode('latin-1')
                    pdf.cell(w=0, h=6, text=clean_line.rstrip(), new_x="LMARGIN", new_y="NEXT")
            
            pdf.output(temp_pdf)
            return index, temp_pdf, True, f"Converted & Ditambahkan: {file_rel_path} -> Text TXT"
        except Exception as e:
            return index, None, False, f"Gagal mengonversi Text TXT {file_rel_path}: {str(e)}"
            
    elif ext == 'xlsx':
        try:
            import openpyxl
            from xhtml2pdf import pisa
            
            # 1. Load Excel and convert sheets to HTML tables
            wb_load = openpyxl.load_workbook(full_path, data_only=True)
            html_sheets = []
            for sheet_name in wb_load.sheetnames:
                sheet = wb_load[sheet_name]
                
                # Cari batas baris dan kolom yang benar-benar berisi data (tidak kosong)
                actual_max_row = 1
                actual_max_col = 1
                for r in range(1, sheet.max_row + 1):
                    for c in range(1, sheet.max_column + 1):
                        val = sheet.cell(row=r, column=c).value
                        if val is not None and str(val).strip() != "":
                            if r > actual_max_row:
                                actual_max_row = r
                            if c > actual_max_col:
                                actual_max_col = c
                
                sheet_html = f"<h2>Sheet: {sheet_name}</h2>"
                sheet_html += "<table>"
                
                for r in range(1, actual_max_row + 1):
                    sheet_html += "<tr>"
                    for c in range(1, actual_max_col + 1):
                        cell_value = sheet.cell(row=r, column=c).value
                        if cell_value is None:
                            cell_value = ""
                        
                        # Gunakan tag th untuk baris pertama (header)
                        if r == 1:
                            sheet_html += f"<th>{cell_value}</th>"
                        else:
                            sheet_html += f"<td>{cell_value}</td>"
                    sheet_html += "</tr>"
                sheet_html += "</table>"
                html_sheets.append(sheet_html)

            all_sheets_html = "<br/>".join(html_sheets)

            styled_html = f"""
            <html>
            <head>
            <meta http-equiv="Content-Type" content="text/html; charset=utf-8">
            <style>
                @page {{
                    size: a4 landscape;
                    margin: 0.5cm;
                }}
                body {{
                    font-family: Helvetica, Arial, sans-serif;
                    font-size: 8pt;
                    color: #333333;
                }}
                h2 {{
                    color: #2c3e50;
                    font-size: 11pt;
                    border-bottom: 0.5pt solid #2c3e50;
                    padding-bottom: 3px;
                    margin-top: 0.4cm;
                    margin-bottom: 0.2cm;
                }}
                table {{
                    border-collapse: collapse;
                    margin-top: 5px;
                    margin-bottom: 15px;
                    width: 100%;
                }}
                th {{
                    border: 0.5pt solid #bdc3c7;
                    padding: 4px;
                    font-size: 7pt;
                    background-color: #f2f2f2;
                    font-weight: bold;
                    text-align: center;
                }}
                td {{
                    border: 0.5pt solid #bdc3c7;
                    padding: 4px;
                    font-size: 7pt;
                    text-align: left;
                }}
                tr:nth-child(even) {{
                    background-color: #f9f9f9;
                }}
            </style>
            </head>
            <body>
                {all_sheets_html}
            </body>
            </html>
            """
            
            # 2. Render HTML to PDF
            with open(temp_pdf, "wb") as result_file:
                pisa_status = pisa.CreatePDF(styled_html, dest=result_file)
                
            if pisa_status.err:
                raise Exception(f"xhtml2pdf error code: {pisa_status.err}")
                
            return index, temp_pdf, True, f"Converted & Ditambahkan: {file_rel_path} -> Excel XLSX"
        except Exception as e:
            return index, None, False, f"Gagal mengonversi Excel XLSX {file_rel_path}: {str(e)}"
            
    else:
        return index, None, False, f"Format file tidak didukung: {file_rel_path}"

if __name__ == "__main__":
    input_file = os.path.join(base_dir, sys.argv[1])
    output_dir = os.path.join(base_dir, "output")

    # Tentukan nama file output
    if len(sys.argv) > 2:
        output_filename = sys.argv[2]
    else:
        output_filename = "hasil_gabungan.pdf"

    output_path = os.path.join(output_dir, output_filename)

    print(f"Base dir: {base_dir}")
    print(f"Input file: {input_file}")

    # Baca input.json
    with open(input_file, 'r') as f:
        all_files = json.load(f)

    print("File input:", all_files)

    # Jalankan konversi secara paralel menggunakan ThreadPoolExecutor
    results = []
    with ThreadPoolExecutor() as executor:
        futures = [executor.submit(process_file, idx, filepath) for idx, filepath in enumerate(all_files)]
        for future in futures:
            try:
                results.append(future.result())
            except Exception as e:
                print("Error eksekusi thread:", str(e))

    # Urutkan hasil kembali sesuai index asli untuk menjaga urutan
    results.sort(key=lambda x: x[0])

    merger = PdfWriter()
    temp_files = []

    # Gabungkan file hasil konversi
    for index, pdf_path, is_temp, log_message in results:
        if log_message:
            print(log_message)
        if pdf_path and os.path.exists(pdf_path):
            try:
                merger.append(pdf_path)
                if is_temp:
                    temp_files.append(pdf_path)
            except Exception as e:
                print(f"Gagal menggabungkan PDF {pdf_path}: {str(e)}")

    # Pastikan folder output ada
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print("Folder output dibuat.")

    # Simpan hasil merger
    try:
        merger.write(output_path)
        merger.close()
        print("Gabungan disimpan di:", output_path)
    except Exception as e:
        print("Gagal menyimpan PDF:", str(e))

    # Hapus file sementara
    for temp in temp_files:
        if os.path.exists(temp):
            try:
                os.remove(temp)
                print("File sementara dihapus:", temp)
            except Exception as e:
                print(f"Gagal menghapus file sementara {temp}: {str(e)}")
